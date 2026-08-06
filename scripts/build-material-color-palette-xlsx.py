# -*- coding: utf-8 -*-
"""從 seed SQL 整理資料，產出可檢視／編輯的配色範例 Excel（雙色＋三色，含英文欄）。"""
import re
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SQL_PATH = ROOT / 'docs' / 'seed-material-color-palette-examples.sql'
ORIGINAL = Path(r'C:\Users\User\Downloads\尼龍布寵物用品色卡 (4) (1).xlsx')
OUT_PROJECT = ROOT / 'docs' / '尼龍布寵物用品色卡-配色範例-更新版.xlsx'
OUT_DOWNLOADS = Path(r'C:\Users\User\Downloads\尼龍布寵物用品色卡-配色範例-更新版-v3.xlsx')

TYPE_EN = {
    '基礎百搭': 'Basic & Versatile',
    '大地戶外': 'Earthy & Outdoor',
    '安全反光': 'Safety & Hi-Vis',
    '運動撞色': 'Sporty Color-Block',
    '清爽甜美': 'Fresh & Sweet',
    '低調質感': 'Understated & Refined',
    '莫蘭迪風': 'Morandi',
    '馬卡龍風': 'Macaron',
    '北歐設計風': 'Nordic Design',
    '精品風': 'Luxury',
    '藝術家': 'Artist-Inspired',
    '設計師/流派': 'Designer & Movement',
    '品牌': 'Brand-Inspired',
}

HDR_FILL = PatternFill('solid', fgColor='2F5233')
HDR_FONT = Font(bold=True, color='FFFFFF')
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hex_to_fill(hex_code):
    h = str(hex_code).strip().upper().lstrip('#')
    if len(h) == 6:
        return PatternFill('solid', fgColor=h)
    return None


def parse_sql():
    text = SQL_PATH.read_text(encoding='utf-8')

    dual = []
    m = re.search(r'WITH new_dual.*?VALUES\s*(.*?)\)\s*INSERT INTO public\.material_color_palettes', text, re.S)
    if not m:
        raise RuntimeError('找不到雙色 VALUES')
    for row in re.findall(
        r"\('([^']*(?:''[^']*)*)',\s*'([^']*(?:''[^']*)*)',\s*'([^']*(?:''[^']*)*)',\s*'(#[0-9A-Fa-f]{6})',\s*'(#[0-9A-Fa-f]{6})',\s*'([^']+)',\s*'\[(\d+),(\d+)\]'::jsonb,\s*'([^']*(?:''[^']*)*)',\s*'([^']*(?:''[^']*)*)',\s*(\d+)\)",
        m.group(1),
    ):
        dual.append({
            'type': row[0].replace("''", "'"),
            'name': row[1].replace("''", "'"),
            'name_en': row[2].replace("''", "'"),
            'hex1': row[3].upper(),
            'hex2': row[4].upper(),
            'ratio_preset': row[5],
            'pct1': int(row[6]),
            'pct2': int(row[7]),
            'note': row[8].replace("''", "'"),
            'note_en': row[9].replace("''", "'"),
        })

    tri = []
    m2 = re.search(r'WITH new_tri.*?VALUES\s*(.*?)\)\s*INSERT INTO public\.material_color_palettes', text, re.S)
    if not m2:
        raise RuntimeError('找不到三色 VALUES')
    for row in re.findall(
        r"\('([^']*(?:''[^']*)*)',\s*'([^']*(?:''[^']*)*)',\s*'([^']*(?:''[^']*)*)',\s*'(#[0-9A-Fa-f]{6})',\s*(\d+),\s*'(#[0-9A-Fa-f]{6})',\s*(\d+),\s*'(#[0-9A-Fa-f]{6})',\s*(\d+),\s*'([^']*(?:''[^']*)*)',\s*'([^']*(?:''[^']*)*)',\s*(\d+)\)",
        m2.group(1),
    ):
        tri.append({
            'type': row[0].replace("''", "'"),
            'name': row[1].replace("''", "'"),
            'name_en': row[2].replace("''", "'"),
            'hex1': row[3].upper(),
            'pct1': int(row[4]),
            'hex2': row[5].upper(),
            'pct2': int(row[6]),
            'hex3': row[7].upper(),
            'pct3': int(row[8]),
            'note': row[9].replace("''", "'"),
            'note_en': row[10].replace("''", "'"),
        })

    return dual, tri


def style_header(ws, headers, widths):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'


def write_dual_sheet(wb, dual_rows):
    if '雙色搭配' in wb.sheetnames:
        del wb['雙色搭配']
    ws = wb.create_sheet('雙色搭配', 1)
    headers = [
        '風格分類', 'Style Category',
        '搭配名稱', 'Palette Name (EN)',
        '主色', 'Primary Color',
        '配色', 'Accent Color',
        '主色色票', 'Primary HEX',
        '配色色票', 'Accent HEX',
        '比重', 'Ratio',
        '備註', 'Notes (EN)',
    ]
    widths = [12, 18, 28, 34, 14, 16, 14, 16, 10, 10, 10, 10, 10, 8, 28, 36]
    style_header(ws, headers, widths)

    for i, r in enumerate(dual_rows, 2):
        ratio = f"{r['pct1']}/{r['pct2']}"
        vals = [
            r['type'], TYPE_EN.get(r['type'], ''),
            r['name'], r['name_en'],
            r['name'].split(' + ')[0] if ' + ' in r['name'] else r['name'],
            r['name_en'].split(' + ')[0] if ' + ' in r['name_en'] else r['name_en'],
            r['name'].split(' + ')[-1] if ' + ' in r['name'] else '',
            r['name_en'].split(' + ')[-1] if ' + ' in r['name_en'] else '',
            r['hex1'], r['hex1'],
            r['hex2'], r['hex2'],
            ratio, ratio,
            r['note'], r['note_en'],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = BORDER
            if col in (9, 11):
                fill = hex_to_fill(val)
                if fill:
                    cell.fill = fill


def write_tri_sheet(wb, tri_rows):
    if '三色搭配' in wb.sheetnames:
        del wb['三色搭配']
    ws = wb.create_sheet('三色搭配', 2)
    headers = [
        '風格分類', 'Style Category',
        '搭配名稱', 'Palette Name (EN)',
        '主色', 'Primary Color',
        '比重1', 'Ratio 1',
        '色票1', 'HEX 1',
        '次色', 'Secondary Color',
        '比重2', 'Ratio 2',
        '色票2', 'HEX 2',
        '點綴色', 'Accent Color',
        '比重3', 'Ratio 3',
        '色票3', 'HEX 3',
        '備註', 'Notes (EN)',
    ]
    widths = [12, 18, 30, 36, 12, 16, 8, 8, 10, 10, 12, 16, 8, 8, 10, 10, 12, 16, 8, 8, 10, 10, 30, 38]
    style_header(ws, headers, widths)

    for i, r in enumerate(tri_rows, 2):
        parts_zh = [p.strip() for p in r['name'].split('+')]
        parts_en = [p.strip() for p in r['name_en'].split('+')]
        while len(parts_zh) < 3:
            parts_zh.append('')
        while len(parts_en) < 3:
            parts_en.append('')
        vals = [
            r['type'], TYPE_EN.get(r['type'], ''),
            r['name'], r['name_en'],
            parts_zh[0], parts_en[0],
            f"{r['pct1']}%", f"{r['pct1']}%",
            r['hex1'], r['hex1'],
            parts_zh[1], parts_en[1],
            f"{r['pct2']}%", f"{r['pct2']}%",
            r['hex2'], r['hex2'],
            parts_zh[2], parts_en[2],
            f"{r['pct3']}%", f"{r['pct3']}%",
            r['hex3'], r['hex3'],
            r['note'], r['note_en'],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = BORDER
            if col in (9, 15, 21):
                fill = hex_to_fill(val)
                if fill:
                    cell.fill = fill


def main():
    dual, tri = parse_sql()
    if ORIGINAL.exists():
        wb = openpyxl.load_workbook(ORIGINAL)
        # 保留原「常用單色」分頁不動
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    write_dual_sheet(wb, dual)
    write_tri_sheet(wb, tri)

    OUT_PROJECT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PROJECT)
    shutil.copy2(OUT_PROJECT, OUT_DOWNLOADS)
    print('DONE')
    print('project:', OUT_PROJECT)
    print('downloads:', OUT_DOWNLOADS)
    print('dual rows:', len(dual), 'tri rows:', len(tri))


if __name__ == '__main__':
    main()

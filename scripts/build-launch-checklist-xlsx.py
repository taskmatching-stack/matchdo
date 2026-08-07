# -*- coding: utf-8 -*-
"""產出 docs/上線前檢查表.xlsx（對照 docs/上線前檢查表.md）。"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / 'docs' / '上線前檢查表.xlsx'

HDR_FILL = PatternFill('solid', fgColor='4A6FA5')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
SEC_FILL = PatternFill('solid', fgColor='E8EEF5')
SEC_FONT = Font(bold=True, size=11)
SUB_FONT = Font(bold=True, color='444444')
NOTE_FILL = PatternFill('solid', fgColor='FFF8E6')
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')

# (區塊, 子區, 項目, 可選)
CHECKLIST_ROWS = [
    ('A', '部署前', 'git status 乾淨；main 已 push 至 origin/main', ''),
    ('A', '部署前', '記錄本次上線 commit（git log -1 --oneline）', ''),
    ('A', '部署前', '必要 env 已在 Cloud Run／Secret 設定（Supabase、BFL、Gemini、PayPal 等）', ''),
    ('A', '部署前', 'node --check server.js 通過', '可選'),
    ('B', '部署', '部署輸出以 Done. 結束', ''),
    ('B', '部署', '流量指向最新 revision（gcloud run services describe …）', ''),
    ('B', '部署', '首頁 https://matchdo.cc/ 可開（非 5xx）', ''),
    ('B', '部署', '靜態資源有更新（BUILD 標記或硬重新整理）', ''),
    ('C', 'Migration', '後台 migration 列表無本次功能依賴仍顯示「未套用」', ''),
    ('D', 'D1 帳號', '登入 /login.html', ''),
    ('D', 'D1 帳號', '/credits.html 餘額顯示正常', ''),
    ('D', 'D1 帳號', '/subscription-plans.html 方案頁可開', ''),
    ('D', 'D2 設計', '/custom-product.html 選分類 + 描述 + 生圖成功', ''),
    ('D', 'D2 設計', '扣點後餘額減少（admin 除外）', ''),
    ('D', 'D2 設計', '我的數位資產 → 設計稿 出現新圖', ''),
    ('D', 'D2 設計', '卡片「履歷」可開；PDF 可下載', ''),
    ('D', 'D2 設計', '（migration 後）履歷含完整模型 prompt（新圖）', ''),
    ('D', 'D3 工具', 'pattern-extract / design-to-physical / scene-sim 至少一項生圖成功', ''),
    ('D', 'D3 工具', '/promo-image/ 情境圖 → 資產庫情境圖分頁', ''),
    ('D', 'D3 工具', '/promo-camera 商攝導演 → 資產庫情境圖分頁', ''),
    ('D', 'D4 材料印花', '/client/material-dual-color.html 生圖 → 材料組合', ''),
    ('D', 'D4 材料印花', '/client/print-asset.html 上傳原圖 → 印花', ''),
    ('D', 'D4 材料印花', '印花 AI 重繪扣點成功', '可選'),
    ('D', 'D5 分享', '設計稿上媒體牆後 /inspiration/user_design/{id} 可開', ''),
    ('D', 'D5 分享', 'OG／分享連結非 404', ''),
    ('E', 'E1 公開', '/vendor-profile.html?id= 作品集、素材庫可見', ''),
    ('E', 'E1 公開', '「用此廠商版型設計」深鏈進設計頁', ''),
    ('E', 'E2 後台', '/client/manufacturer-portfolio.html 上傳或編輯展示案例', ''),
    ('E', 'E2 後台', '/client/manufacturer-materials.html 上傳或編輯素材（付費帳）', ''),
    ('E', 'E2 後台', '素材 AI 重繪或放大預覽至少一項可用', ''),
    ('E', 'E3 B線', '/client/industry-suppliers.html 可瀏覽', '需展示案例'),
    ('E', 'E3 B線', '匯入一筆 → my-supplier-references 可見', '需展示案例'),
    ('E', 'E4 Embed', '素材庫取得 iframe 嵌入碼', '付費廠商'),
    ('E', 'E4 Embed', '/embed/simulator.html 訪客試做生圖（廠商扣點）', '付費廠商'),
    ('E', 'E4 Embed', '/client/embed-design-records.html 有紀錄；履歷可開', '付費廠商'),
    ('F', '供應商', '/client/supplier-catalog-manage.html 可上架（付費供應商）', ''),
    ('F', '供應商', '/client/industry-supplier-dashboard.html 引用紀錄可載入', ''),
    ('G', 'G1 營運', '/admin/platform-stats.html 各 tab；表尾總計正常', ''),
    ('G', 'G1 營運', '情境圖統計可區分情境圖頁 / 商攝導演', ''),
    ('G', 'G1 營運', '/admin/generation-records.html 列表、詳情 modal 正常', ''),
    ('G', 'G1 營運', '勾選 2 筆 → 匯出選取 ZIP 可下載且可解開', ''),
    ('G', 'G2 字典', '/admin/custom-categories.html 可開', ''),
    ('G', 'G2 字典', '/admin/promo-scene-templates.html 可開', ''),
    ('G', 'G2 字典', '/admin/material-color-palettes.html 可開', '需 migration'),
    ('G', 'G2 字典', '官方版型庫 ?official_platform=1&manage=1 可開', ''),
    ('G', 'G3 會員', '/admin/membership.html 可查看／手動調點', ''),
    ('G', 'G3 會員', '/admin/user-management.html 可搜尋測試帳號', ''),
    ('H', '金流權限', '免費帳號上傳素材／供應商品 → 403 或頁內提示（限制 A）', ''),
    ('H', '金流權限', '無展示案例製造商匯入 B 線 → 阻擋（限制 B）', ''),
    ('H', '金流權限', 'PayPal 儲值或訂閱走完 → 點數到帳', ''),
    ('H', '金流權限', '生圖失敗時未異常扣點（抽測）', ''),
    ('I', 'SEO', '/official-templates/ 真列表（非 301 進設計 tab）', ''),
    ('I', 'SEO', '/vendor-styles/ 真列表可開', ''),
    ('I', 'SEO', '/sitemap.xml 可開且子 sitemap 非空', ''),
    ('I', 'SEO', '/client/* 個人後台 noindex', ''),
    ('I', 'SEO', 'promo-camera?embed=design noindex；/promo-camera 可索引', ''),
    ('J', '多語系', '首頁或工具 ?lang=en 導覽為英文', '可選'),
    ('J', '多語系', '官方配色／情境模板英文名稱有顯示', '可選'),
]

MIGRATION_ROWS = [
    ('provenance-resume-fields', '生圖履歷：扣點 FK、composed prompt、衍生鏈', '必跑'),
    ('user-material-combo-generations', '材料組合資產庫表', '依現網'),
    ('user-print-generations', '印花資產庫表', '依現網'),
    ('user-asset-library-category', '資產庫分類欄', '依現網'),
    ('user-material-presets', '材料組合常用文字', '依現網'),
    ('material-color-palettes', '官方配色範例表', '依現網'),
    ('material-color-palette-ratios', '配色比重欄', '依現網'),
    ('material-color-palette-i18n', '配色多語系', '依現網'),
    ('material-color-palette-notes', '配色備註', '依現網'),
    ('material-color-palette-examples-seed', '配色種子資料', '可選'),
]

DEPLOY_LINES = [
    'gcloud config set account taskmatching@gmail.com',
    'gcloud config set project matchdo',
    '',
    'cd ~/matchdo && git fetch origin main && git reset --hard origin/main && ( gcloud run deploy matchdo --source . --region=asia-northeast1 --allow-unauthenticated --clear-base-image && gcloud run services update-traffic matchdo --region=asia-northeast1 --to-latest ) 2>&1 | grep --line-buffered -v -E \'Regional Access Boundary|taskmatchlng\'',
    '',
    'gcloud run services describe matchdo --region=asia-northeast1 --format=\'yaml(spec.traffic,status.latestReadyRevisionName)\'',
]

SIGNOFF_ROWS = [
    ('上線日期', ''),
    ('部署 commit', ''),
    ('執行者', ''),
    ('Migration 已跑', ''),
    ('阻擋問題（若有）', ''),
    ('結論', '☐ 可上線  ☐ 延後'),
]


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build():
    wb = openpyxl.Workbook()

    # --- 檢查表 ---
    ws = wb.active
    ws.title = '檢查表'
    meta = [
        'MatchDO 上線前檢查表',
        '適用：matchdo.cc 正式／Beta | 基準 commit：82b4052 | 對照：docs/上線前檢查表.md',
        '完成欄請填：是／否／N/A；備註欄記錄問題或截圖連結',
    ]
    for i, line in enumerate(meta, 1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = Font(bold=(i == 1), size=12 if i == 1 else 10)
        if i == 1:
            cell.fill = PatternFill('solid', fgColor='D6E4F0')

    headers = ['區塊', '子區', '檢查項目', '完成', '備註', '可選']
    hr = 5
    for col, h in enumerate(headers, 1):
        ws.cell(row=hr, column=col, value=h)
    style_header(ws, hr, len(headers))

    r = hr + 1
    prev_block = None
    for block, sub, item, opt in CHECKLIST_ROWS:
        if block != prev_block:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            label = f'區塊 {block}'
            ws.cell(row=r, column=1, value=label).font = SEC_FONT
            ws.cell(row=r, column=1).fill = SEC_FILL
            r += 1
            prev_block = block
        ws.cell(row=r, column=1, value=block)
        ws.cell(row=r, column=2, value=sub)
        ws.cell(row=r, column=3, value=item)
        ws.cell(row=r, column=4, value='☐')
        ws.cell(row=r, column=6, value=opt)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = WRAP
        r += 1

    ws.freeze_panes = 'A6'
    set_col_widths(ws, [8, 14, 52, 8, 24, 12])

    # --- Migration ---
    wm = wb.create_sheet('Migration')
    wm.append(['migration id', '用途', '優先級', '完成', '備註'])
    style_header(wm, 1, 5)
    for mid, desc, pri in MIGRATION_ROWS:
        wm.append([mid, desc, pri, '☐', ''])
    for row in wm.iter_rows(min_row=2, max_row=wm.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = BORDER
            cell.alignment = WRAP
    set_col_widths(wm, [28, 40, 10, 8, 30])
    wm.freeze_panes = 'A2'

    # --- 部署指令 ---
    wd = wb.create_sheet('部署指令')
    wd.merge_cells('A1:B1')
    wd['A1'] = 'Cloud Shell 部署（見 deploy-matchdo-push-and-deploy.md §3.1）'
    wd['A1'].font = Font(bold=True, size=12)
    for i, line in enumerate(DEPLOY_LINES, 3):
        wd.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        c = wd.cell(row=i, column=1, value=line)
        c.alignment = WRAP
        c.font = Font(name='Consolas', size=10)
    set_col_widths(wd, [100, 20])

    # --- 簽核 ---
    ws2 = wb.create_sheet('簽核')
    ws2.append(['項目', '值'])
    style_header(ws2, 1, 2)
    for k, v in SIGNOFF_ROWS:
        ws2.append([k, v])
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = WRAP
    set_col_widths(ws2, [22, 50])

    # --- Backlog ---
    wb.create_sheet('Backlog')
    wb['Backlog']['A1'] = '已知不擋上線（上線後再做）'
    wb['Backlog']['A1'].font = Font(bold=True)
    backlog = [
        '生圖履歷舊資料 backfill',
        '設計風向（測試中）',
        '廠商內容英文自動翻譯 UI',
        'Embed CAPTCHA／域名白名單',
        '攝影 App Store／IAP',
        'SEO Phase D 內容債',
    ]
    for i, b in enumerate(backlog, 2):
        wb['Backlog'][f'A{i}'] = b
    wb['Backlog'].column_dimensions['A'].width = 40

    wb.save(OUT)
    print('Wrote', OUT)


if __name__ == '__main__':
    build()
